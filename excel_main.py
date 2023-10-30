# -*- coding: utf-8 -*-
import sys
import os
sys.path.append(os.path.dirname(__file__))
sys.path.append('../excel')

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
# from openpyxl.formula.Parser import parse
# from openpyxl.formula.evaluator import Evaluator
import openpyxl

                  

from database.repository import Repository
wb = load_workbook(filename='nine.xlsx')
print(wb.sheetnames)
for i in wb.sheetnames:
    print(type(i))
    print(i)
print(type(wb.sheetnames))
# sheet_range1 = wb['认证情况']
# print(sheet_range1['E74'].value)
# sheet_range1['E74'] = '你好'


# # sheet_ranges = wb['统计']
# # column = get_column_letter(7)  # 获取列号对应的字母，此例中为G
# # formula = '=IFNA(VLOOKUP(C2, 认证情况!$B:$N, 3, FALSE), "未认证")'
# # cell = sheet_ranges[f'{column}2']
# # cell.value = formula

# wb.save('hello_1.xlsx')




def write(num_list):                                
    # num_list = [1,2,3,4,5,6]           
    # num_list = ['餐饮','WDhespsw','汉恩斯披萨王(金山万达)','汉恩斯披萨王(金山万达)','已认证','汉恩斯披萨王','已认证','微信用户']           
    bg = load_workbook("nine.xlsx")      	# 应先将excel文件放入到工作目录下 
    sheet = bg["认证情况"]   
    # num_list为二维数组
    # i为行，j为列
    for i in range(1, len(num_list)+1):
        for j in range(1,len(num_list[i-1])+1):
            # 表示将num_list列表的第0个元素的第0个数据，写入到excel表格的第2行第一列				
            sheet.cell(i+1 , j, num_list[i - 1][j-1])					
    bg.save("nine_1018.xlsx")   
    print('写入成功')


""" 读取sheet所有的公式 """
def read_format():
    # 打开文本文件以写入模式
    output_file = open('output.txt', encoding="utf-8", mode='w')

    # 遍历所有Sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        output_file.write(f"Sheet Name: {sheet_name}\n")

        # 遍历每一行
        for row in sheet.iter_rows():
            # 遍历每个单元格
            for cell in row:
                if cell.data_type == 'f' and cell.value is not None:
                    output_file.write(f"Cell Formula: {cell.value}\n")
    # 关闭文本文件
    output_file.close()



def read_user_login_status():
    pass

def read_self_evaluation_situation():
    pass

def read_self_evaluation_situation():
    pass
def read_self_evaluation_situation():
    pass
def read_authentication_situation():
    pass


if __name__ == '__main__':
    rep = Repository()
    # data = rep.read_authentication_situation()
    # write(data)